from os.path import isfile
from string import ascii_uppercase

from openpyxl import load_workbook

class Get:

    def __init__(self, filePath:str):
        self.path = filePath

    def _createLocalId(self, letterInt:int, colInt:int=1):

        letter = ascii_uppercase[letterInt]
        col = str(colInt)

        return letter + col

    def read(self):

        wb = load_workbook(self.path)
        ws = wb.active
        
        # get the headers
        headers = []
        headerint = 0
        while True:
            ident = self._createLocalId(headerint)

            val = ws[ident].value
            if val is None:
                break
            
            headers.append(
                val
            )

            headerint += 1

        rows = []
        rowint = 2
        k = True
        while k:
            letterint = 0
            row = {}
            for header in headers:
                ident = self._createLocalId(letterint, rowint)
                val = ws[ident].value

                if val == None:
                    return rows # it returns 

                # row.append(ws[ident].value)
                row[header] = ws[ident].value
                
                letterint += 1
            rows.append(row)
            rowint += 1

                
